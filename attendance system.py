import tkinter as tk
from tkinter import messagebox
import cv2
from pyzbar.pyzbar import decode
from PIL import Image, ImageTk
import face_recognition
from openpyxl import load_workbook
from datetime import datetime
import os
import numpy as np
import threading

# Function to scan barcodes and return the roll number
def barcode_scanner():
    cap = cv2.VideoCapture(0)
    if not cap.isOpened():
        print("Error: Could not open camera.")
        return None

    cap.set(3, 640)  # Set the width
    cap.set(4, 480)  # Set the height

    print("Starting barcode scanner. Press 'q' to exit.")
    barcode_data = None

    while True:
        ret, frame = cap.read()
        if not ret:
            print("Error: Could not read frame.")
            break

        gray_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        barcodes = decode(gray_frame)

        if barcodes:
            for barcode in barcodes:
                barcode_data = barcode.data.decode('utf-8')
                return barcode_data

        cv2.imshow('Barcode Scanner', frame)
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    cap.release()
    cv2.destroyAllWindows()
    return barcode_data

# Function to display the student's image by roll number
def display_image_by_roll_number(roll_number, folder_path):
    possible_files = [f"{roll_number}.png", f"{roll_number}.jpg"]
    roll_folder_path = os.path.join(folder_path, str(roll_number))
    
    if os.path.isdir(roll_folder_path):
        for file_name in possible_files:
            image_path = os.path.join(roll_folder_path, file_name)
            if os.path.isfile(image_path):
                return image_path

    return None

# Function to retrieve student details from a text file
def get_student_details(roll_number, folder_path):
    details_file = os.path.join(folder_path, str(roll_number), f"{roll_number}.doc")
    if os.path.isfile(details_file):
        with open(details_file, 'r') as file:
            details = file.read()
        return details
    return "Student details not found."

# Function to mark attendance in Excel
def mark_attendance(roll_number):
    file_path = "attendance.xlsx"
    
    if not os.path.exists(file_path) or not file_path.endswith('.xlsx'):
        print("Attendance file not found or invalid format.")
        return
    
    try:
        wb = load_workbook(file_path)
        sheet = wb.active
        found = False
        for row in sheet.iter_rows(min_row=2, values_only=False):
            cell = row[0]
            if cell.value == roll_number:
                found = True
                cell_attendance = row[1]
                cell_attendance.value = "Present"
                cell_timestamp = row[2]
                cell_timestamp.value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                break

        if not found:
            sheet.append([roll_number, "Present", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])

        wb.save(file_path)
        print(f"Attendance marked for roll number: {roll_number}")
    except Exception as e:
        print(f"Error saving the attendance file: {e}")

# Live face recognition function
def live_face_recognition(known_image_path, roll_number, tolerance=0.5):
    known_image = face_recognition.load_image_file(known_image_path)
    known_encoding = face_recognition.face_encodings(known_image)[0]

    cap = cv2.VideoCapture(0)
    attendance_marked = False
    print("Starting live face recognition. Press 'q' to exit.")

    while True:
        ret, frame = cap.read()
        if not ret:
            break

        rgb_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        face_locations = face_recognition.face_locations(rgb_frame)
        face_encodings = face_recognition.face_encodings(rgb_frame, face_locations)

        for (top, right, bottom, left), face_encoding in zip(face_locations, face_encodings):
            face_distances = face_recognition.face_distance([known_encoding], face_encoding)
            matches = face_recognition.compare_faces([known_encoding], face_encoding, tolerance=tolerance)
            name = "Unknown"

            if matches[0] and face_distances[0] < tolerance and not attendance_marked:
                name = roll_number
                mark_attendance(roll_number)
                attendance_marked = True
                print(f"Attendance marked for roll number: {roll_number}")
                break

            cv2.rectangle(frame, (left, top), (right, bottom), (0, 255, 0), 2)
            cv2.putText(frame, name, (left, top - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 255, 0), 2)

        if attendance_marked or cv2.waitKey(1) & 0xFF == ord('q'):
            break

        cv2.imshow('Live Face Recognition', frame)

    cap.release()
    cv2.destroyAllWindows()

# Tkinter frontend code
class AttendanceSystemApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Smart Attendance System")
        self.root.geometry("600x500")
        self.root.configure(bg="#2b2b2b")

        # Title Label
        title_frame = tk.Frame(self.root, bg="#333")
        title_frame.pack(fill="x", pady=20)
        tk.Label(title_frame, text="Smart Attendance System", font=("Helvetica", 24, "bold"), bg="#333", fg="#fafafa").pack(pady=10)

        # Main Frame
        main_frame = tk.Frame(self.root, bg="#2b2b2b")
        main_frame.pack(pady=20)

        # Scan Barcode Button
        self.scan_button = tk.Button(main_frame, text="📷 Scan Barcode", font=("Helvetica", 16, "bold"), bg="#ff8c00", fg="#fff", command=self.start_barcode_scanner, width=20)
        self.scan_button.grid(row=0, column=0, padx=10, pady=10)

        # Label for displaying student image and details
        self.image_label = tk.Label(main_frame, bg="#2b2b2b")
        self.image_label.grid(row=1, column=0, pady=10)
        self.details_label = tk.Label(main_frame, text="", font=("Helvetica", 12), bg="#2b2b2b", fg="#fafafa", wraplength=500)
        self.details_label.grid(row=2, column=0, pady=10)

        self.roll_number = None
        self.image_path = None

    def start_barcode_scanner(self):
        self.roll_number = barcode_scanner()
        if self.roll_number:
            self.display_student_info()
            self.start_face_recognition()
        else:
            messagebox.showerror("Error", "No barcode detected.")

    def display_student_info(self):
        folder_path = "PHOTO"
        self.image_path = display_image_by_roll_number(self.roll_number, folder_path)
        
        if self.image_path:
            img = Image.open(self.image_path)
            img = img.resize((100, 200), Image.LANCZOS)
            tk_img = ImageTk.PhotoImage(img)
            self.image_label.configure(image=tk_img)
            self.image_label.image = tk_img
        else:
            messagebox.showerror("Error", "Image not found for the given roll number.")

        details = get_student_details(self.roll_number, folder_path)
        self.details_label.config(text=details)

    def start_face_recognition(self):
        if self.roll_number and self.image_path:
            live_face_recognition(self.image_path, self.roll_number)

# Run the application
if __name__ == "__main__":
    root = tk.Tk()
    app = AttendanceSystemApp(root)
    root.mainloop()
